#!/usr/bin/env python3
"""
cloudcp_fallback_test.py — CloudCp transfer-fallback test harness.

Drives the fallback test matrix documented in ``plan_cp_fallback.md``. For each
selected case it:

  1. uploads the datagen spec to the Bryck and materialises the dataset
     (``datagen --spec <spec>``) under ``/bryck/cloudcp_fallback/<tier>``;
  2. patches the ``TEST`` fault-injection block (+ ``TRANSFER.HI_PERF_OPT`` /
     ``FALLBACK_ENABLED``) of ``/etc/bryck/bryckcloud/config.json`` and restarts
     ``bcloud.service`` so the new values take effect;
  3. initiates the transfer through the Bryck REST API (reusing the
     ``bryckclient-cli`` modules) and captures the transfer id;
  4. polls the transfer to a terminal state and — WHILE it is IN_PROGRESS —
     snapshots the live internal ``*.txt.lst`` and ``cloudcp_retry_<id>_*.lst``
     files server-side (cloudcp deletes them on completion), then pulls them
     down;
  5. downloads + parses the transfer report (``FALLBACK_OK`` / ``SUCCESS`` rows);
  6. evaluates the case verdict against its expectation.

Case logic: with faults injected, a transfer that COMPLETES means the fallback
works; a transfer that FAILS means the fallback does not. The negative cases
disable ``FALLBACK_ENABLED`` and expect the transfer to FAIL.

Every host command and API call is recorded as an ordered "step" so the final
JSON + HTML reports show exactly what ran.

Usage
-----
    python3 cloudcp_fallback_test.py --list
    python3 cloudcp_fallback_test.py --all
    python3 cloudcp_fallback_test.py --one FB-U-07
    python3 cloudcp_fallback_test.py --one FB-U-03,FB-U-07,FB-D-02
    python3 cloudcp_fallback_test.py --from FB-U-01 --to FB-U-10
    python3 cloudcp_fallback_test.py --negative
    python3 cloudcp_fallback_test.py --negative-case FB-N-01
    python3 cloudcp_fallback_test.py --all --dry-run        # print the plan only

The runner drives the Bryck remotely over SSH + REST (paramiko / requests),
reusing ``login.json`` and ``cloud_ops.json`` from the ``bryckclient-cli``
directory. Passwordless ``sudo -n`` is required on the Bryck for the config
write and the service restart (same as the other client runners).
"""
from __future__ import annotations

import argparse
import csv
import datetime as _dt
import html
import json
import logging
import sys
import time
import zipfile
from dataclasses import dataclass
from pathlib import Path
from typing import Any

HERE = Path(__file__).resolve().parent
DEFAULT_CLI_DIR = HERE / "bryckclient-cli"
DEFAULT_SPEC_DIR = HERE / "spec_files"
DEFAULT_OUT_DIR = HERE / "runs"

LOG = logging.getLogger("fallback_test")

# ---- Host-side defaults (override via CLI) ---------------------------------
DEF_DATAGEN = "/home/bryck/rperiyas/datagen"
DEF_CONFIG = "/etc/bryck/bryckcloud/config.json"
DEF_SERVICE = "bcloud.service"
DEF_TRANSFER_LOGS = "/opt/bryck/bryckapi/downloads/cloud_transfer_logs"
DEF_ENDPOINT = "https://10.10.10.103:9000"
DEF_SRC_BASE = "/bryck/cloudcp_fallback"
DEF_BUCKET_BASE = "s3://omicron/fallback"
DEF_DL_BASE = "/bryck/cloudcp_fallback_dl"
DEF_REMOTE_SPEC_DIR = "/tmp/fb_specs"
DEF_REMOTE_CAPTURE_DIR = "/tmp/fb_capture"

DEF_POLL_INTERVAL = 10
DEF_POLL_TIMEOUT = 3600
DEF_CONFIGURE_TIMEOUT = 60
DEF_PAUSED_TIMEOUT = 300

# Terminal transfer states (mirrors the client scripts).
TERMINAL_SUCCESS_STATE = "COMPLETED"
TERMINAL_FAILURE_STATES = {"FAILED", "STOPPED", "CANCELLED", "PAUSED_TIMEOUT"}
TERMINAL_STATES = {TERMINAL_SUCCESS_STATE} | TERMINAL_FAILURE_STATES

# Per-file report statuses that count as "transferred".
TERMINAL_SUCCESS_ROWS = {"SUCCESS", "SKIPPED", "FALLBACK_OK", "MP_OK"}

STDOUT_CLIP = 4000  # max chars of captured stdout/stderr kept per step


# =============================================================================
# Fault profiles
# =============================================================================

@dataclass(frozen=True)
class Profile:
    name: str
    fail_pct: int
    crash_pct: int
    desc: str


PROFILES: dict[str, Profile] = {
    "F0": Profile("F0", 0, 0, "Control / baseline — no fault; plain success path."),
    "F1": Profile("F1", 10, 0, "Light per-file failures; fallback retries a few records."),
    "F2": Profile("F2", 50, 0, "Half the records fail on the primary path."),
    "F3": Profile("F3", 100, 0, "Every record fails the primary path — fallback carries all."),
    "F4": Profile("F4", 0, 10, "Occasional worker crash (abort)."),
    "F5": Profile("F5", 0, 50, "Frequent worker crashes."),
    "F6": Profile("F6", 0, 100, "Every record crashes the primary worker."),
    "F7": Profile("F7", 100, 100, "Saturation extreme — fail + crash everywhere."),
}

DEFAULT_FAULT_SEED = 1337


# =============================================================================
# Dataset registry
# =============================================================================

@dataclass(frozen=True)
class Dataset:
    key: str
    spec: str            # spec filename under spec_files/
    tier_dir: str        # leaf dir name = the spec's root leaf
    checksum_stable: bool = False
    heavy: bool = False


DATASETS: dict[str, Dataset] = {
    "zero":    Dataset("zero", "01_zero_byte.yaml", "zero_byte"),
    "tiny":    Dataset("tiny", "02_tiny_files.yaml", "tiny_files"),
    "small":   Dataset("small", "03_small_files.yaml", "small_files"),
    "medium":  Dataset("medium", "04_medium_files.yaml", "medium_files"),
    "large":   Dataset("large", "05_large_files.yaml", "large_files", heavy=True),
    "sparse":  Dataset("sparse", "06_sparse_files.yaml", "sparse_files"),
    "fill":    Dataset("fill", "07_fill_files.yaml", "fill_files", checksum_stable=True),
    "deep":    Dataset("deep", "08_deep_tree.yaml", "deep_tree"),
    "unicode": Dataset("unicode", "09_unicode_names.yaml", "unicode_names"),
    "special": Dataset("special", "10_special_char_names.yaml", "special_char_names"),
    "mixed":   Dataset("mixed", "11_mixed_realistic.yaml", "mixed_realistic"),
    "scale":   Dataset("scale", "12_tiny_2million.yaml", "tiny_2million", heavy=True),
}


# =============================================================================
# Expectation + Case
# =============================================================================

# Expectation "kind" drives verdict evaluation.
#   fallback_ok : faulted profile, fallback ON  -> COMPLETED + retry lists + all files.
#   control     : F0, fallback ON               -> COMPLETED + all files, no retry lists.
#   must_fail   : faulted profile, fallback OFF -> terminal FAILURE.
#   clean_fail  : invalid target                -> terminal FAILURE / init error, no crash.
Direction = str  # "upload" | "download"


@dataclass
class Case:
    cid: str
    group: str                 # UPLOAD | DOWNLOAD | MIN-ACCEPT | NEGATIVE
    direction: Direction
    dataset: str               # key into DATASETS
    profile: str               # key into PROFILES
    hi_perf: bool
    fallback_enabled: bool
    expect: str                # fallback_ok | control | must_fail | clean_fail
    desc: str
    invalid_bucket: bool = False

    @property
    def ds(self) -> Dataset:
        return DATASETS[self.dataset]

    @property
    def prof(self) -> Profile:
        return PROFILES[self.profile]


# =============================================================================
# Case catalog (mirrors plan_cp_fallback.md)
# =============================================================================

def build_catalog() -> list[Case]:
    cases: list[Case] = []

    def add(cid, group, direction, dataset, profile, hi_perf, fb, expect, desc,
            invalid_bucket=False):
        cases.append(Case(cid, group, direction, dataset, profile, hi_perf, fb,
                          expect, desc, invalid_bucket))

    # ---- §8 Upload matrix (HI_PERF_OPT=True, FALLBACK_ENABLED=True) ----------
    add("FB-U-01", "UPLOAD", "upload", "zero", "F0", True, True, "control",
        "Control upload of zero-byte files with no fault — must succeed cleanly with no fallback.")
    add("FB-U-02", "UPLOAD", "upload", "zero", "F3", True, True, "fallback_ok",
        "All zero-byte (empty-object) records fail the primary path and are carried by the fallback.")
    add("FB-U-03", "UPLOAD", "upload", "tiny", "F1", True, True, "fallback_ok",
        "Light 10% failures across a high count of tiny files; a few records retried via fallback.")
    add("FB-U-04", "UPLOAD", "upload", "tiny", "F3", True, True, "fallback_ok",
        "Every tiny record fails the primary path — fallback must carry the entire tiny workload.")
    add("FB-U-05", "UPLOAD", "upload", "tiny", "F6", True, True, "fallback_ok",
        "Every tiny record crashes the primary worker (abort); fallback recovers all records.")
    add("FB-U-06", "UPLOAD", "upload", "small", "F2", True, True, "fallback_ok",
        "Half the small files fail while straddling the 8 MiB multipart cutoff; fallback retries them.")
    add("FB-U-07", "UPLOAD", "upload", "small", "F3", True, True, "fallback_ok",
        "All small records fail the primary path; fallback carries single-part and multipart alike.")
    add("FB-U-08", "UPLOAD", "upload", "small", "F7", True, True, "fallback_ok",
        "Saturation (100% fail + 100% crash) on the multipart boundary; fallback must still complete.")
    add("FB-U-09", "UPLOAD", "upload", "medium", "F1", True, True, "fallback_ok",
        "Light failures on multi-chunk multipart medium files; a few chunk records retried via fallback.")
    add("FB-U-10", "UPLOAD", "upload", "medium", "F3", True, True, "fallback_ok",
        "All medium multipart uploads fail the primary path; fallback carries every multipart object.")
    add("FB-U-11", "UPLOAD", "upload", "medium", "F6", True, True, "fallback_ok",
        "Every medium multipart worker crashes; fallback recovers all multipart records.")
    add("FB-U-12", "UPLOAD", "upload", "large", "F1", True, True, "fallback_ok",
        "Light failures on large (GiB) multipart; fallback retries the few failed large objects.")
    add("FB-U-13", "UPLOAD", "upload", "large", "F3", True, True, "fallback_ok",
        "All large records fail the primary path (long-running); fallback carries the whole large tier.")
    add("FB-U-14", "UPLOAD", "upload", "large", "F7", True, True, "fallback_ok",
        "Saturation on large multipart; fallback must still drive the transfer to completion.")
    add("FB-U-15", "UPLOAD", "upload", "sparse", "F3", True, True, "fallback_ok",
        "Sparse-content files all fail the primary path; fallback handles logical-vs-physical sizes.")
    add("FB-U-16", "UPLOAD", "upload", "fill", "F3", True, True, "fallback_ok",
        "Deterministic fill files all fail the primary path; fallback data is byte-exact (checksum verify).")
    add("FB-U-17", "UPLOAD", "upload", "deep", "F3", True, True, "fallback_ok",
        "Deeply nested / long-key files fail the primary path; keys survive the fallback retry intact.")
    add("FB-U-18", "UPLOAD", "upload", "unicode", "F3", True, True, "fallback_ok",
        "Unicode / emoji / CJK filenames all fail the primary path; names round-trip through fallback.")
    add("FB-U-19", "UPLOAD", "upload", "special", "F3", True, True, "fallback_ok",
        "ASCII special-char / space filenames all fail; fallback preserves the object keys verbatim.")
    add("FB-U-20", "UPLOAD", "upload", "mixed", "F2", True, True, "fallback_ok",
        "Realistic mixed workload with half the records faulted; fallback retries across all tiers.")
    add("FB-U-21", "UPLOAD", "upload", "mixed", "F7", True, True, "fallback_ok",
        "Saturation on the realistic mixed workload; fallback must complete the whole mix.")
    add("FB-U-22", "UPLOAD", "upload", "scale", "F1", True, True, "fallback_ok",
        "Scale: ~2M tiny files with light failures; stresses fallback retry-list handling at volume.")

    # ---- §9 Download matrix -------------------------------------------------
    add("FB-D-01", "DOWNLOAD", "download", "small", "F0", True, True, "control",
        "Control download of small files with no fault — must succeed cleanly with no fallback.")
    add("FB-D-02", "DOWNLOAD", "download", "small", "F3", True, True, "fallback_ok",
        "All small download records fail the primary path; fallback carries the download to completion.")
    add("FB-D-03", "DOWNLOAD", "download", "tiny", "F1", True, True, "fallback_ok",
        "Light failures downloading many tiny objects; a few records retried via download fallback.")
    add("FB-D-04", "DOWNLOAD", "download", "medium", "F3", True, True, "fallback_ok",
        "All medium multipart downloads fail the primary path; fallback carries the multipart download.")
    add("FB-D-05", "DOWNLOAD", "download", "large", "F6", True, True, "fallback_ok",
        "Every large download worker crashes; fallback recovers the large-object download.")
    add("FB-D-06", "DOWNLOAD", "download", "fill", "F3", True, True, "fallback_ok",
        "All fill-file downloads fail the primary path; downloaded bytes are byte-exact via fallback.")
    add("FB-D-07", "DOWNLOAD", "download", "mixed", "F7", True, True, "fallback_ok",
        "Saturation download of the realistic mix; fallback must complete the whole download.")

    # ---- §10 Minimum acceptance (HI_PERF_OPT=False) -------------------------
    add("FB-HP-01", "MIN-ACCEPT", "upload", "zero", "F3", False, True, "fallback_ok",
        "Zero tier with HI_PERF_OPT off; all records fail the primary path, fallback carries them.")
    add("FB-HP-02", "MIN-ACCEPT", "upload", "tiny", "F3", False, True, "fallback_ok",
        "Tiny tier with HI_PERF_OPT off; all records fail the primary path, fallback carries them.")
    add("FB-HP-03", "MIN-ACCEPT", "upload", "small", "F3", False, True, "fallback_ok",
        "Small tier with HI_PERF_OPT off; all records fail the primary path, fallback carries them.")
    add("FB-HP-04", "MIN-ACCEPT", "upload", "medium", "F3", False, True, "fallback_ok",
        "Medium tier with HI_PERF_OPT off; all records fail the primary path, fallback carries them.")
    add("FB-HP-05", "MIN-ACCEPT", "upload", "large", "F3", False, True, "fallback_ok",
        "Large tier with HI_PERF_OPT off; all records fail the primary path, fallback carries them.")

    # ---- §11 Negative scenarios ---------------------------------------------
    add("FB-N-01", "NEGATIVE", "upload", "small", "F1", True, False, "must_fail",
        "FALLBACK_ENABLED=False with light failures — faulted records are NOT recovered; transfer must FAIL.")
    add("FB-N-02", "NEGATIVE", "upload", "tiny", "F3", True, False, "must_fail",
        "FALLBACK_ENABLED=False with all records failing — no fallback to carry them; transfer must FAIL.")
    add("FB-N-03", "NEGATIVE", "upload", "medium", "F6", True, False, "must_fail",
        "FALLBACK_ENABLED=False with all-crash — no recovery path; transfer must FAIL.")
    add("FB-N-04", "NEGATIVE", "upload", "large", "F7", True, False, "must_fail",
        "FALLBACK_ENABLED=False under saturation — no fallback; large transfer must FAIL.")
    add("FB-N-05", "NEGATIVE", "upload", "small", "F7", True, True, "fallback_ok",
        "Positive counterpart to FB-N-01: same saturation but FALLBACK_ENABLED=True must COMPLETE.")
    add("FB-N-06", "NEGATIVE", "upload", "zero", "F0", True, True, "clean_fail",
        "Invalid (nonexistent) destination bucket with no fault — must fail cleanly, no crash/hang.",
        invalid_bucket=True)
    add("FB-N-07", "NEGATIVE", "download", "small", "F3", True, False, "must_fail",
        "FALLBACK_ENABLED=False download with all records failing — download must FAIL.")

    return cases


# =============================================================================
# Step recorder
# =============================================================================

def _now() -> str:
    return _dt.datetime.now().isoformat(timespec="seconds")


def _clip(text: str | None) -> str:
    if not text:
        return ""
    text = text.strip()
    if len(text) > STDOUT_CLIP:
        return text[:STDOUT_CLIP] + f"\n… [clipped, {len(text)} chars total]"
    return text


class Recorder:
    """Collects ordered steps (command + result) for one case."""

    def __init__(self, cid: str):
        self.cid = cid
        self.steps: list[dict] = []

    def add(self, name: str, kind: str, command: str, *, rc: int | None = None,
            ok: bool | None = None, stdout: str = "", stderr: str = "",
            detail: str = "") -> dict:
        step = {
            "seq": len(self.steps) + 1,
            "ts": _now(),
            "name": name,
            "kind": kind,            # ssh | api | local | plan
            "command": command,
            "rc": rc,
            "ok": ok,
            "stdout": _clip(stdout),
            "stderr": _clip(stderr),
            "detail": detail,
        }
        self.steps.append(step)
        level = logging.INFO if ok in (True, None) else logging.WARNING
        LOG.log(level, "[%s] %s :: %s", self.cid, name, detail or command)
        return step


# =============================================================================
# Remote host wrapper (SSH) — records every command
# =============================================================================

class RemoteHost:
    """Thin SSH facade that records each command as a step (or plans it on dry-run)."""

    def __init__(self, ssh, dry_run: bool):
        self.ssh = ssh          # ssh_runner.SshRunner | None (None on dry-run)
        self.dry_run = dry_run

    def run(self, rec: Recorder, name: str, cmd: str, *, timeout: int = 120,
            check: bool = False) -> tuple[int, str, str]:
        if self.dry_run or self.ssh is None:
            rec.add(name, "plan", cmd, detail="(dry-run) would run over SSH")
            return 0, "", ""
        rc, out, err = self.ssh.run(cmd, timeout=timeout)
        ok = (rc == 0)
        rec.add(name, "ssh", cmd, rc=rc, ok=ok, stdout=out, stderr=err,
                detail=f"rc={rc}")
        if check and rc != 0:
            raise RuntimeError(f"remote command failed (rc={rc}): {cmd}\n{err.strip()}")
        return rc, out, err

    def put(self, rec: Recorder, name: str, local: str, remote: str) -> None:
        cmd = f"sftp put {local} -> {remote}"
        if self.dry_run or self.ssh is None:
            rec.add(name, "plan", cmd, detail="(dry-run) would SFTP put")
            return
        self.ssh.put(local, remote)
        rec.add(name, "ssh", cmd, ok=True, detail="uploaded")

    def get(self, rec: Recorder, name: str, remote: str, local: str) -> bool:
        cmd = f"sftp get {remote} -> {local}"
        if self.dry_run or self.ssh is None:
            rec.add(name, "plan", cmd, detail="(dry-run) would SFTP get")
            return False
        try:
            self.ssh.get(remote, local)
            rec.add(name, "ssh", cmd, ok=True, detail="downloaded")
            return True
        except Exception as exc:  # noqa: BLE001
            rec.add(name, "ssh", cmd, ok=False, detail=f"get failed: {exc}")
            return False


# =============================================================================
# Config patching
# =============================================================================

def read_remote_config(host: RemoteHost, rec: Recorder, config_path: str) -> dict:
    rc, out, _ = host.run(rec, "read config", f"cat {config_path}", check=False)
    if host.dry_run:
        return {}
    if rc != 0 or not out.strip():
        raise RuntimeError(f"could not read {config_path}")
    return json.loads(out)


def apply_fault_config(host: RemoteHost, rec: Recorder, cfg: dict, case: Case,
                       config_path: str, service: str, seed: int,
                       out_dir: Path) -> None:
    """Patch TEST + TRANSFER blocks, push to the Bryck, restart the service."""
    prof = case.prof
    patched = json.loads(json.dumps(cfg)) if cfg else _template_config()
    test = patched.setdefault("TEST", {})
    test["SIM_TRANSFER"] = False
    test["FAULT_FAIL_PERCENT"] = prof.fail_pct
    test["FAULT_CRASH_PERCENT"] = prof.crash_pct
    test["FAULT_CRASH_MODE"] = "abort"
    test["FAULT_SEED"] = seed

    transfer = patched.setdefault("TRANSFER", {})
    transfer["HI_PERF_OPT"] = "True" if case.hi_perf else "False"
    transfer["FALLBACK_ENABLED"] = "True" if case.fallback_enabled else "False"

    summary = (
        f"FAIL%={prof.fail_pct} CRASH%={prof.crash_pct} SEED={seed} "
        f"CRASH_MODE=abort HI_PERF_OPT={transfer['HI_PERF_OPT']} "
        f"FALLBACK_ENABLED={transfer['FALLBACK_ENABLED']}"
    )

    if host.dry_run:
        rec.add("patch config", "plan",
                f"edit {config_path}: {summary}",
                detail="(dry-run) would rewrite TEST + TRANSFER and restart the service")
        rec.add("restart service", "plan",
                f"sudo -n systemctl restart {service}",
                detail="(dry-run)")
        return

    local_tmp = out_dir / f"config_{case.cid}.json"
    local_tmp.write_text(json.dumps(patched, indent=2), encoding="utf-8")
    remote_tmp = f"/tmp/fb_config_{case.cid}.json"
    host.put(rec, "upload patched config", str(local_tmp), remote_tmp)
    host.run(rec, "install config (sudo)",
             f"sudo -n cp {remote_tmp} {config_path} && rm -f {remote_tmp}",
             check=True)
    rec.add("config applied", "local", summary, ok=True, detail=summary)
    host.run(rec, "restart service (sudo)",
             f"sudo -n systemctl restart {service}", timeout=120, check=True)
    # Give the service a moment to come up before hitting the API.
    host.run(rec, "await service", f"systemctl is-active {service} || true", check=False)


def _template_config() -> dict:
    return {"TEST": {}, "TRANSFER": {}}


def restore_config(host: RemoteHost, rec: Recorder, backup_remote: str,
                   config_path: str, service: str) -> None:
    if host.dry_run:
        rec.add("restore config", "plan",
                f"sudo -n cp {backup_remote} {config_path}; restart {service}",
                detail="(dry-run)")
        return
    host.run(rec, "restore original config",
             f"sudo -n cp {backup_remote} {config_path}", check=False)
    host.run(rec, "restart service (restore)",
             f"sudo -n systemctl restart {service}", check=False)


# =============================================================================
# Cloud transfer via the bryckclient-cli API modules
# =============================================================================

def _extract_transfer_id(result: Any) -> str:
    if isinstance(result, (str, int)):
        return str(result)
    if isinstance(result, dict):
        for key in ("transfer_id", "id"):
            if result.get(key):
                return str(result[key])
    raise RuntimeError(f"no transfer_id in response result: {result!r}")


def _status_entry(api, transfer_id: str) -> dict | None:
    resp = api.get_cloud_transfer_status(transfer_id)
    if resp is None:
        return None
    # Detect auth failure responses that slipped through.
    if getattr(resp, "status_code", 200) in (401, 403):
        LOG.warning("status poll got HTTP %s — token may have expired", resp.status_code)
        return None
    try:
        body = resp.json()
    except ValueError:
        return None
    result = body.get("result")
    if isinstance(result, list):
        for item in result:
            if isinstance(item, dict) and item:
                return item
        return None
    if isinstance(result, dict) and result:
        return result
    return None


def configure_cloud_provider(api, rec: Recorder, creds: dict) -> None:
    rec.add("configure cloud", "api",
            f"POST /api/bcloud/config bcloud_type={creds['cloud_type']}",
            detail="configuring provider credentials")
    resp = api.configure_cloud(
        bcloud_type=creds["cloud_type"],
        username=creds.get("access_key_id"),
        keyid=creds.get("secret_access_key"),
        region=creds.get("region"),
    )
    # 409 CONFLICT means the provider is already configured -> not an error.
    if getattr(resp, "status_code", None) == 409:
        rec.steps[-1]["ok"] = True
        rec.steps[-1]["detail"] = "already configured (HTTP 409 ignored)"
        return
    ok = resp is not None
    deadline = time.time() + DEF_CONFIGURE_TIMEOUT
    listed = False
    while ok and time.time() < deadline:
        lresp = api.get_cloud_config_list()
        try:
            configs = lresp.json().get("result", []) if lresp is not None else []
        except (ValueError, AttributeError):
            configs = []
        for entry in configs or []:
            etype = str((entry or {}).get("bcloud_type") or (entry or {}).get("cloud_type") or "").lower()
            if etype == creds["cloud_type"].lower():
                listed = True
                break
        if listed:
            break
        time.sleep(2)
    rec.steps[-1]["ok"] = ok and listed
    rec.steps[-1]["detail"] = "provider configured" if listed else "configure not confirmed"


def _bucket_root(bucket_base: str) -> str:
    """s3://aditya/fallback -> s3://aditya (bucket root for a recursive wipe)."""
    if "://" not in bucket_base:
        return bucket_base
    scheme, _, path = bucket_base.partition("://")
    bucket = path.split("/", 1)[0]
    return f"{scheme}://{bucket}"


def _server_error_message(resp) -> str:
    """Best-effort extraction of the server-side error reason from a Response."""
    if resp is None:
        return ""
    try:
        payload = resp.json()
    except (ValueError, AttributeError):
        return (getattr(resp, "text", "") or "")[:300]
    if isinstance(payload, dict):
        err = payload.get("error")
        if isinstance(err, dict) and err.get("message"):
            return str(err["message"])
        for key in ("message", "detail", "result"):
            if payload.get(key):
                return str(payload[key])
    return (getattr(resp, "text", "") or "")[:300]


def initiate_transfer(api, rec: Recorder, cloud_type: str, src: str, dst: str) -> str:
    rec.add("initiate transfer", "api",
            f"POST /api/bcloud/transfer src={src} dst={dst}",
            detail="starting transfer")
    resp = api.initiate_cloud_transfer(cloud_type, src, dst)
    if resp is None or getattr(resp, "status_code", 0) != 200:
        code = getattr(resp, "status_code", "None")
        reason = _server_error_message(resp)
        detail = f"initiate failed (HTTP {code})"
        if reason:
            detail = f"{detail}: {reason}"
        rec.steps[-1]["ok"] = False
        rec.steps[-1]["detail"] = detail
        raise TransferInitError(detail)
    tid = _extract_transfer_id(resp.json().get("result"))
    rec.steps[-1]["ok"] = True
    rec.steps[-1]["detail"] = f"transfer_id={tid}"
    return tid


class TransferInitError(RuntimeError):
    pass


def poll_transfer(api, host: RemoteHost, rec: Recorder, transfer_id: str,
                  log_dir_base: str, capture_dir: Path, interval: int,
                  timeout: int, session=None) -> tuple[str, dict]:
    """Poll to a terminal state; snapshot live *.lst files while IN_PROGRESS."""
    remote_log_dir = f"{log_dir_base}/cloud_transfer_{transfer_id}"
    remote_snap = f"{DEF_REMOTE_CAPTURE_DIR}/{transfer_id}"
    host.run(rec, "prep capture dir", f"mkdir -p {remote_snap}", check=False)

    # Start a background watcher that copies .lst files every second.
    # This avoids missing short-lived files between poll intervals.
    # The loop is bounded (timeout + slack) so it self-terminates even if the
    # harness dies before the explicit kill below — no orphaned watchers.
    watcher_pid_file = f"/tmp/fb_watcher_{transfer_id}.pid"
    watcher_max_secs = int(timeout) + 120
    watcher_cmd = (
        f"nohup bash -c '"
        f"for _ in $(seq 1 {watcher_max_secs}); do "
        f"  cp -a {remote_log_dir}/*.txt.lst {remote_snap}/ 2>/dev/null; "
        f"  cp -a {remote_log_dir}/cloudcp_retry_{transfer_id}_*.lst {remote_snap}/ 2>/dev/null; "
        f"  sleep 1; "
        f"done"
        f"' >/dev/null 2>&1 & echo $! > {watcher_pid_file}"
    )
    host.run(rec, "start .lst watcher", watcher_cmd, check=False)

    deadline = time.time() + timeout
    last_state = ""
    polls = 0
    paused_since: float | None = None
    consecutive_none = 0
    try:
        while time.time() < deadline:
            polls += 1

            # Proactively refresh the token before polling to avoid mid-poll expiry.
            if session is not None:
                session.ensure_token()

            entry = _status_entry(api, transfer_id)

            # Handle consecutive failed status calls (likely auth issues).
            if entry is None:
                consecutive_none += 1
                if consecutive_none >= 3 and session is not None:
                    LOG.warning("3 consecutive failed status calls — forcing re-login")
                    try:
                        session.login(silent=True)
                    except Exception:  # noqa: BLE001
                        pass
                    consecutive_none = 0
                rec.add(f"status (poll {polls})", "api",
                        f"GET /api/bcloud/status_transfer transfer_id={transfer_id}",
                        ok=False, detail="no response (possible token expiry)")
                time.sleep(interval)
                continue
            consecutive_none = 0

            state = str((entry or {}).get("state") or "").upper()
            pct = (entry or {}).get("percent_completed")
            rec.add(f"status (poll {polls})", "api",
                    f"GET /api/bcloud/status_transfer transfer_id={transfer_id}",
                    ok=True, detail=f"state={state or '?'} percent={pct}")
            last_state = state or last_state
            if state in TERMINAL_STATES:
                break

            # Abort if stuck in PAUSED for longer than the paused timeout.
            if state == "PAUSED":
                if paused_since is None:
                    paused_since = time.time()
                elif time.time() - paused_since >= DEF_PAUSED_TIMEOUT:
                    rec.add("paused timeout", "local", "",
                            ok=False,
                            detail=f"transfer stuck in PAUSED for >{DEF_PAUSED_TIMEOUT}s — aborting poll")
                    last_state = "PAUSED_TIMEOUT"
                    break
            else:
                paused_since = None

            time.sleep(interval)
    finally:
        # Always stop the watcher, even if the poll loop raises / is interrupted.
        host.run(rec, "stop .lst watcher",
                 f"kill $(cat {watcher_pid_file} 2>/dev/null) 2>/dev/null; rm -f {watcher_pid_file}",
                 check=False)

    # Do one final capture pass.
    host.run(rec, "final .lst capture",
             f"cp -a {remote_log_dir}/*.txt.lst {remote_snap}/ 2>/dev/null; "
             f"cp -a {remote_log_dir}/cloudcp_retry_{transfer_id}_*.lst {remote_snap}/ 2>/dev/null; "
             f"ls -1 {remote_snap} 2>/dev/null | wc -l",
             check=False)

    # Post-completion: .lst files get renamed to .txt.lst.done and batches
    # move to completed/; these persist after the transfer finishes.
    host.run(rec, "capture .lst.done files",
             f"cp -a {remote_log_dir}/*.txt.lst.done {remote_snap}/ 2>/dev/null; "
             f"ls -1 {remote_log_dir}/*.txt.lst.done 2>/dev/null | wc -l",
             check=False)
    host.run(rec, "capture completed batches",
             f"mkdir -p {remote_snap}/completed 2>/dev/null; "
             f"cp -a {remote_log_dir}/completed/* {remote_snap}/completed/ 2>/dev/null; "
             f"ls -1 {remote_log_dir}/completed/ 2>/dev/null | wc -l",
             check=False)

    # Pull the accumulated snapshot down.
    lst_count = _collect_snapshot(host, rec, transfer_id, remote_snap, capture_dir)

    # Count post-completion evidence separately.
    done_count = _count_remote_files(host, rec, f"{remote_log_dir}/*.txt.lst.done")
    completed_count = _count_remote_files(host, rec, f"{remote_log_dir}/completed/*")
    retry_count = _count_remote_files(
        host, rec, f"{remote_log_dir}/cloudcp_retry_{transfer_id}_*.lst*")

    return last_state, {
        "polls": polls,
        "captured_lst": lst_count,
        "done_lst": done_count,
        "completed_batches": completed_count,
        "retry_lst": retry_count,
        "remote_snapshot": remote_snap,
    }


def _count_remote_files(host: RemoteHost, rec: Recorder, glob: str) -> int:
    """Count files matching a remote glob (0 on dry-run or error)."""
    if host.dry_run:
        return 0
    rc, out, _ = host.run(rec, "count remote files",
                          f"ls -1 {glob} 2>/dev/null | wc -l", check=False)
    try:
        return int(out.strip().splitlines()[-1]) if out.strip() else 0
    except (ValueError, IndexError):
        return 0


def _collect_snapshot(host: RemoteHost, rec: Recorder, transfer_id: str,
                      remote_snap: str, capture_dir: Path) -> int:
    capture_dir.mkdir(parents=True, exist_ok=True)
    tgz_remote = f"/tmp/fb_capture_{transfer_id}.tgz"
    host.run(rec, "archive captured .lst",
             f"tar czf {tgz_remote} -C {remote_snap} . 2>/dev/null; "
             f"ls -1 {remote_snap} 2>/dev/null | wc -l",
             check=False)
    local_tgz = capture_dir / f"live_lst_{transfer_id}.tgz"
    got = host.get(rec, "download captured .lst", tgz_remote, str(local_tgz))
    count = 0
    if got and local_tgz.is_file():
        try:
            import tarfile
            with tarfile.open(local_tgz, "r:gz") as tf:
                members = [m for m in tf.getmembers() if m.isfile()]
                count = len(members)
                _safe_extract(tf, capture_dir / f"live_lst_{transfer_id}")
        except Exception as exc:  # noqa: BLE001
            rec.add("extract .lst", "local", str(local_tgz), ok=False,
                    detail=f"extract failed: {exc}")
    rec.add("captured .lst summary", "local", "",
            ok=True, detail=f"{count} live list file(s) captured")
    return count


def _safe_extract(tf, dest: Path) -> None:
    dest.mkdir(parents=True, exist_ok=True)
    base = dest.resolve()
    for member in tf.getmembers():
        target = (dest / member.name).resolve()
        if not str(target).startswith(str(base)):
            continue  # skip path traversal entries
        tf.extract(member, dest)


# =============================================================================
# System logs — fallback worker evidence
# =============================================================================

def check_fallback_in_logs(api, rec: Recorder, transfer_id: str) -> dict:
    """Query GET /api/config/getlogs and look for fallback worker messages."""
    rec.add("fetch system logs", "api",
            "GET /api/config/getlogs?cursor=Today",
            detail="checking for fallback worker log entries")
    resp = api.get_logs(cursor="Today")
    if resp is None or not getattr(resp, "ok", False):
        rec.steps[-1]["ok"] = False
        rec.steps[-1]["detail"] = "getlogs request failed"
        return {"fallback_started": False, "fallback_done": False,
                "fb_transferred": 0, "fb_failed": 0}

    try:
        entries = resp.json().get("result", [])
    except (ValueError, AttributeError):
        entries = []

    tid_str = str(transfer_id)
    fb_started = False
    fb_done = False
    fb_transferred = 0
    fb_failed = 0

    for entry in entries or []:
        msg = str((entry or {}).get("message") or "")
        if f"Fallback worker started: transfer_id={tid_str}" in msg:
            fb_started = True
        if "Fallback worker done:" in msg and fb_started:
            fb_done = True
            # Parse "transferred=N failed=M"
            for part in msg.split():
                if part.startswith("transferred="):
                    try:
                        fb_transferred = int(part.split("=", 1)[1])
                    except ValueError:
                        pass
                if part.startswith("failed="):
                    try:
                        fb_failed = int(part.split("=", 1)[1])
                    except ValueError:
                        pass

    detail = (f"fallback_started={fb_started} fallback_done={fb_done} "
              f"transferred={fb_transferred} failed={fb_failed}")
    rec.steps[-1]["ok"] = True
    rec.steps[-1]["detail"] = detail
    return {"fallback_started": fb_started, "fallback_done": fb_done,
            "fb_transferred": fb_transferred, "fb_failed": fb_failed}


# =============================================================================
# Report download + parse
# =============================================================================

def download_and_parse_report(api, rec: Recorder, transfer_id: str,
                              run_dir: Path) -> dict:
    rec.add("download report", "api",
            f"GET /api/download?name=cloud_log&type={transfer_id}",
            detail="fetching report zip")
    resp = api.download_cloud_transfer_log(transfer_id)
    zip_path = run_dir / f"cloud_transfer_report_{transfer_id}.zip"
    if resp is None or not getattr(resp, "ok", False):
        rec.steps[-1]["ok"] = False
        rec.steps[-1]["detail"] = "report not available"
        return {"available": False}
    written = 0
    try:
        with open(zip_path, "wb") as fh:
            for chunk in resp.iter_content(chunk_size=8192):
                if chunk:
                    fh.write(chunk)
                    written += len(chunk)
    finally:
        resp.close()
    rec.steps[-1]["ok"] = written > 0
    rec.steps[-1]["detail"] = f"saved {written} bytes -> {zip_path.name}"
    if written == 0:
        return {"available": False}

    extract_dir = run_dir / f"report_{transfer_id}"
    try:
        with zipfile.ZipFile(zip_path) as zf:
            zf.extractall(extract_dir)
    except zipfile.BadZipFile:
        rec.add("parse report", "local", str(zip_path), ok=False,
                detail="downloaded file is not a valid zip")
        return {"available": False}

    metrics = _parse_report_dir(extract_dir, transfer_id)
    rec.add("parse report", "local", str(extract_dir), ok=True,
            detail=(f"rows={metrics['total_rows']} success={metrics['success_rows']} "
                    f"fallback_ok={metrics['fallback_ok']} failed={metrics['failed_rows']} "
                    f"retried={metrics['retried_rows']} mp_ok={metrics['mp_ok_rows']}"))
    metrics["available"] = True
    return metrics


def _parse_report_dir(extract_dir: Path, transfer_id: str = "") -> dict:
    total = success = fallback_ok = failed = 0
    retried = mp_ok = 0
    files_seen: set[str] = set()

    # --- Parse CSV files ---
    for csv_path in extract_dir.rglob("*.csv"):
        _parse_tabular_rows(csv_path, files_seen,
                            _count_cb_factory(locals_ref := {"t": 0, "s": 0, "fb": 0,
                                                             "f": 0, "r": 0, "mp": 0}))
    total += locals_ref["t"]; success += locals_ref["s"]
    fallback_ok += locals_ref["fb"]; failed += locals_ref["f"]
    retried += locals_ref["r"]; mp_ok += locals_ref["mp"]

    # --- Parse XLSX (transfer_report_<id>.xlsx) ---
    for xlsx_path in extract_dir.rglob("*.xlsx"):
        _parse_xlsx_report(xlsx_path, files_seen,
                           _count_cb_factory(x_ref := {"t": 0, "s": 0, "fb": 0,
                                                       "f": 0, "r": 0, "mp": 0}))
        total += x_ref["t"]; success += x_ref["s"]
        fallback_ok += x_ref["fb"]; failed += x_ref["f"]
        retried += x_ref["r"]; mp_ok += x_ref["mp"]

    # --- Parse final_report.json ---
    for jr in extract_dir.rglob("final_report.json"):
        _parse_final_report_json(jr, files_seen,
                                 _count_cb_factory(j_ref := {"t": 0, "s": 0, "fb": 0,
                                                             "f": 0, "r": 0, "mp": 0}))
        total += j_ref["t"]; success += j_ref["s"]
        fallback_ok += j_ref["fb"]; failed += j_ref["f"]
        retried += j_ref["r"]; mp_ok += j_ref["mp"]

    return {"total_rows": total, "success_rows": success,
            "fallback_ok": fallback_ok, "failed_rows": failed,
            "retried_rows": retried, "mp_ok_rows": mp_ok}


def _count_cb_factory(ref: dict):
    """Return a callback that tallies a single report row into ref."""
    def cb(status: str, attempt: int, path: str, files_seen: set[str]):
        if path and path in files_seen:
            return
        if path:
            files_seen.add(path)
        ref["t"] += 1
        if status in TERMINAL_SUCCESS_ROWS:
            ref["s"] += 1
        if status in ("FALLBACK_OK", "MP_OK"):
            ref["fb"] += 1
        if status == "MP_OK":
            ref["mp"] += 1
        if status not in TERMINAL_SUCCESS_ROWS:
            ref["f"] += 1
        if attempt >= 2:
            ref["r"] += 1
    return cb


def _parse_tabular_rows(csv_path: Path, files_seen: set[str], cb) -> None:
    """Parse a CSV report file, calling cb(status, attempt, path) per row."""
    try:
        with open(csv_path, newline="", encoding="utf-8", errors="replace") as fh:
            reader = csv.DictReader(fh)
            if reader.fieldnames is None:
                return
            lower_fields = [(f or "").strip().lower() for f in reader.fieldnames]
            if "status" not in lower_fields:
                return
            key_map = {lf: orig for lf, orig in zip(lower_fields, reader.fieldnames)}
            status_key = key_map.get("status")
            path_key = key_map.get("local_path") or key_map.get("absolutefilepath") or key_map.get("s3path")
            attempt_key = key_map.get("attempt")
            for row in reader:
                status = (row.get(status_key) or "").strip().upper()
                lp = (row.get(path_key) or "") if path_key else ""
                try:
                    attempt = int(row.get(attempt_key) or "1") if attempt_key else 1
                except ValueError:
                    attempt = 1
                cb(status, attempt, lp, files_seen)
    except OSError:
        pass


def _parse_xlsx_report(xlsx_path: Path, files_seen: set[str], cb) -> None:
    """Parse an xlsx transfer report for status/attempt columns."""
    try:
        import openpyxl
    except ImportError:
        return
    try:
        wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    except Exception:  # noqa: BLE001
        return
    try:
        for ws in wb.worksheets:
            headers = [str(c.value or "").strip().lower() for c in next(ws.iter_rows(max_row=1))]
            col = {h: i for i, h in enumerate(headers)}
            status_idx = col.get("status")
            if status_idx is None:
                continue
            attempt_idx = col.get("attempt")
            path_idx = col.get("local_path") or col.get("s3path")
            for row in ws.iter_rows(min_row=2, values_only=True):
                status = str(row[status_idx] or "").strip().upper()
                lp = str(row[path_idx] or "") if path_idx is not None else ""
                try:
                    attempt = int(row[attempt_idx]) if attempt_idx is not None and row[attempt_idx] else 1
                except (ValueError, TypeError):
                    attempt = 1
                cb(status, attempt, lp, files_seen)
    finally:
        wb.close()


def _parse_final_report_json(json_path: Path, files_seen: set[str], cb) -> None:
    """Parse final_report.json for status/attempt fields."""
    try:
        data = json.loads(json_path.read_text(encoding="utf-8"))
    except (OSError, ValueError):
        return
    entries = data if isinstance(data, list) else data.get("entries", data.get("result", []))
    if not isinstance(entries, list):
        return
    for entry in entries:
        if not isinstance(entry, dict):
            continue
        status = str(entry.get("status") or "").strip().upper()
        lp = str(entry.get("local_path") or entry.get("s3path") or "")
        try:
            attempt = int(entry.get("attempt", 1))
        except (ValueError, TypeError):
            attempt = 1
        cb(status, attempt, lp, files_seen)


# =============================================================================
# Verdict evaluation
# =============================================================================

def evaluate(case: Case, state: str, capture: dict, report: dict,
             expected_files: int | None) -> tuple[str, list[str]]:
    """Return (verdict, reasons). verdict in PASS | FAIL | WARN."""
    reasons: list[str] = []
    completed = state == TERMINAL_SUCCESS_STATE
    failed = state in TERMINAL_FAILURE_STATES
    fb_ok = report.get("fallback_ok", 0)
    retry_lists = capture.get("captured_lst", 0)
    success_rows = report.get("success_rows", 0)

    if case.expect == "fallback_ok":
        if not completed:
            return "FAIL", [f"expected COMPLETED, got {state or 'no-terminal-state'}"]
        # Primary evidence: system logs confirm fallback worker started.
        logs_started = capture.get("logs_fallback_started", False)
        # Secondary evidence: report rows with attempt >= 2 or MP_OK status.
        retried = report.get("retried_rows", 0)
        mp_ok = report.get("mp_ok_rows", 0)
        # Tertiary: captured .lst/.done/completed artifacts.
        done_lst = capture.get("done_lst", 0)
        completed_batches = capture.get("completed_batches", 0)
        retry_files = capture.get("retry_lst", 0)

        has_fallback_evidence = (
            logs_started or retried > 0 or mp_ok > 0
            or retry_lists > 0 or done_lst > 0 or completed_batches > 0
            or retry_files > 0 or fb_ok > 0
        )
        if not has_fallback_evidence:
            reasons.append("no fallback evidence (logs / attempt>=2 / MP_OK / .lst artifacts)")
        if expected_files and report.get("available") and success_rows < expected_files:
            reasons.append(f"transferred {success_rows} < expected {expected_files} files")
        return ("PASS" if has_fallback_evidence and not reasons else
                ("PASS" if not reasons else "WARN")), reasons

    if case.expect == "control":
        if not completed:
            return "FAIL", [f"control expected COMPLETED, got {state or 'no-terminal-state'}"]
        if expected_files and report.get("available") and success_rows < expected_files:
            reasons.append(f"transferred {success_rows} < expected {expected_files} files")
        if report.get("available") and report.get("fallback_ok", 0) > 0:
            reasons.append("unexpected FALLBACK_OK rows in a no-fault control run")
        return ("PASS" if not reasons else "WARN"), reasons

    if case.expect == "must_fail":
        if failed:
            return "PASS", [f"transfer failed as expected (state={state})"]
        if completed:
            return "FAIL", ["transfer COMPLETED with fallback disabled — faults not applied or fallback still active"]
        return "FAIL", [f"expected terminal FAILURE, got {state or 'no-terminal-state'}"]

    if case.expect == "clean_fail":
        if failed:
            return "PASS", [f"failed cleanly as expected (state={state})"]
        if completed:
            return "FAIL", ["transfer COMPLETED against an invalid target"]
        return "WARN", [f"no clean terminal failure observed (state={state or 'none'})"]

    return "WARN", [f"unknown expectation kind {case.expect!r}"]


# =============================================================================
# Case runner
# =============================================================================

class Runner:
    def __init__(self, args, api, host: RemoteHost, creds: dict, session=None):
        self.args = args
        self.api = api
        self.host = host
        self.creds = creds
        self.session = session
        self.out_dir = Path(args.out_dir)
        self.spec_dir = Path(args.spec_dir)
        self.seeded_tiers: set[str] = set()
        self.config_backup_remote = "/tmp/fb_config_backup.json"
        self._backed_up = False

    # -- one-time original config backup ------------------------------------
    def backup_config_once(self, rec: Recorder) -> None:
        if self._backed_up or self.host.dry_run:
            if self.host.dry_run:
                rec.add("backup config", "plan",
                        f"cp {self.args.config} {self.config_backup_remote}",
                        detail="(dry-run)")
            return
        self.host.run(rec, "backup original config",
                      f"test -f {self.config_backup_remote} || "
                      f"cp {self.args.config} {self.config_backup_remote}",
                      check=False)
        self._backed_up = True

    def ensure_dataset(self, rec: Recorder, ds: Dataset) -> int | None:
        """Upload the spec and run datagen; return generated file count (or None)."""
        src = f"{self.args.src_base}/{ds.tier_dir}"
        if self.args.skip_datagen:
            rec.add("skip datagen", "local", f"reuse {src}", detail="--skip-datagen")
            return self._count_files(rec, src)
        local_spec = self.spec_dir / ds.spec
        remote_spec = f"{DEF_REMOTE_SPEC_DIR}/{ds.spec}"
        self.host.run(rec, "prep remote spec dir",
                      f"mkdir -p {DEF_REMOTE_SPEC_DIR}", check=False)
        self.host.put(rec, "upload spec", str(local_spec), remote_spec)
        self.host.run(rec, "datagen",
                      f"{self.args.datagen} --spec {remote_spec}",
                      timeout=self.args.poll_timeout, check=True)
        return self._count_files(rec, src)

    def _count_files(self, rec: Recorder, path: str) -> int | None:
        rc, out, _ = self.host.run(rec, "count source files",
                                   f"find {path} -type f 2>/dev/null | wc -l",
                                   check=False)
        if self.host.dry_run:
            return None
        try:
            return int(out.strip().splitlines()[-1]) if out.strip() else 0
        except (ValueError, IndexError):
            return None

    def seed_bucket_if_needed(self, rec: Recorder, case: Case) -> bool:
        """For download cases, ensure objects exist via a clean F0 upload.

        Returns True if the destination bucket should now hold the objects
        (or seeding is not required); False if the seeding upload failed so the
        caller can abort the download instead of hitting a confusing 409.
        """
        ds = case.ds
        if case.direction != "download":
            return True
        if self.args.skip_seed:
            rec.add("skip seed", "local", "", detail="--skip-seed: reuse existing objects")
            return True
        # The per-case cleanup wipes the whole bucket, so a cached seed is only
        # valid when cleanup is disabled (objects persist across cases).
        if ds.tier_dir in self.seeded_tiers and self.args.skip_cleanup:
            rec.add("seed cached", "local", "", detail=f"{ds.tier_dir} already seeded this run")
            return True
        seed_case = Case(f"{case.cid}-seed", case.group, "upload", case.dataset,
                         "F0", True, True, "control", "clean seed upload for download")
        src = f"{self.args.src_base}/{ds.tier_dir}"
        dst = f"{self.args.bucket_base}/{ds.tier_dir}"
        rec.add("seed bucket", "local",
                f"clean F0 upload {src} -> {dst}", detail="seeding objects for download")
        self._configure_and_apply(rec, seed_case)
        if self.host.dry_run:
            self.seeded_tiers.add(ds.tier_dir)
            return True

        configure_cloud_provider(self.api, rec, self.creds)
        seed_ok = False
        try:
            tid = initiate_transfer(self.api, rec, self.creds["cloud_type"], src, dst)
            state, _ = poll_transfer(self.api, self.host, rec, tid,
                                     self.args.transfer_logs_dir,
                                     self.out_dir / f"{case.cid}-seed",
                                     self.args.poll_interval, self.args.poll_timeout,
                                     session=self.session)
            seed_ok = (state == TERMINAL_SUCCESS_STATE)
            rec.add("seed result", "local", "", ok=seed_ok,
                    detail=f"seed upload state={state}")
        except TransferInitError as exc:
            rec.add("seed result", "local", "", ok=False, detail=str(exc))

        # Verify the objects actually landed before relying on them downstream.
        if seed_ok:
            _, out, _ = self.host.run(
                rec, "verify seed objects",
                f"aws s3 ls --recursive {dst} --endpoint-url {self.args.endpoint_url} "
                f"2>/dev/null | wc -l", check=False)
            try:
                objs = int(out.strip().splitlines()[-1]) if out.strip() else 0
            except (ValueError, IndexError):
                objs = 0
            if objs == 0:
                seed_ok = False
                rec.add("seed verify", "local", "", ok=False,
                        detail=f"no objects found under {dst} after seed upload")
            else:
                rec.add("seed verify", "local", "", ok=True,
                        detail=f"{objs} object(s) present under {dst}")

        if seed_ok:
            self.seeded_tiers.add(ds.tier_dir)
        return seed_ok


    def _configure_and_apply(self, rec: Recorder, case: Case) -> None:
        cfg = read_remote_config(self.host, rec, self.args.config)
        apply_fault_config(self.host, rec, cfg, case, self.args.config,
                           self.args.service, self.args.seed, self.out_dir)

    def run_case(self, case: Case) -> dict:
        rec = Recorder(case.cid)
        run_dir = self.out_dir / case.cid
        run_dir.mkdir(parents=True, exist_ok=True)
        started = _now()

        ds = case.ds
        src = f"{self.args.src_base}/{ds.tier_dir}"
        if case.invalid_bucket:
            dst = f"{self.args.bucket_base}/__nonexistent__/{ds.tier_dir}"
        else:
            dst = f"{self.args.bucket_base}/{ds.tier_dir}"
        dl_dst = f"{self.args.dl_base}/{ds.tier_dir}"

        rec.add("case", "local", "",
                detail=(f"{case.cid} [{case.group}] {case.direction} "
                        f"dataset={ds.key} profile={case.profile} "
                        f"HI_PERF_OPT={case.hi_perf} FALLBACK_ENABLED={case.fallback_enabled}"))
        rec.add("description", "local", "", detail=case.desc)

        verdict, reasons = "ERROR", []
        state = ""
        capture: dict = {}
        report: dict = {"available": False}
        expected_files: int | None = None
        transfer_id: str | None = None

        try:
            self.backup_config_once(rec)
            expected_files = self.ensure_dataset(rec, ds)

            # Download cases need objects present first: datagen -> upload seed.
            seed_ok = self.seed_bucket_if_needed(rec, case)
            if case.direction == "download" and not seed_ok:
                raise RuntimeError(
                    f"seed upload failed — source objects missing at {dst}; "
                    "cannot run the download case")

            # Apply this case's fault profile + toggles.
            self._configure_and_apply(rec, case)

            if self.host.dry_run:
                # Plan the transfer + verification without executing.
                if case.direction == "upload":
                    rec.add("initiate transfer", "plan",
                            f"POST /api/bcloud/transfer src={src} dst={dst}",
                            detail="(dry-run)")
                else:
                    rec.add("initiate transfer", "plan",
                            f"POST /api/bcloud/transfer src={dst} dst={dl_dst}",
                            detail="(dry-run)")
                rec.add("poll status", "plan",
                        "GET /api/bcloud/status_transfer (loop) + snapshot .lst",
                        detail="(dry-run)")
                rec.add("download+parse report", "plan",
                        "GET /api/download?name=cloud_log", detail="(dry-run)")
                verdict, reasons = "PLANNED", ["dry-run: no execution"]
            else:
                # Ensure token is fresh before starting API-heavy operations.
                if self.session is not None:
                    self.session.ensure_token()
                configure_cloud_provider(self.api, rec, self.creds)
                if case.direction == "upload":
                    t_src, t_dst = src, dst
                else:
                    t_src, t_dst = dst, dl_dst
                    # cloudcp skips objects whose original local copy still
                    # exists, so clear the upload source and hand the download a
                    # fresh, separate target dir (distinct from the upload src).
                    src_rm = f"rm -rf {src}" if ds.tier_dir else "true"
                    self.host.run(rec, "clear source before download",
                                  src_rm, check=False)
                    self.host.run(rec, "prep download target",
                                  f"mkdir -p {dl_dst}", check=False)
                try:
                    transfer_id = initiate_transfer(
                        self.api, rec, self.creds["cloud_type"], t_src, t_dst)
                except TransferInitError as exc:
                    # For clean_fail (invalid bucket) an init error IS the pass.
                    if case.expect == "clean_fail":
                        verdict, reasons = "PASS", [f"initiate rejected as expected: {exc}"]
                    else:
                        verdict, reasons = "FAIL", [f"initiate failed: {exc}"]
                    transfer_id = None

                if transfer_id is not None:
                    state, capture = poll_transfer(
                        self.api, self.host, rec, transfer_id,
                        self.args.transfer_logs_dir, run_dir / "live_lst",
                        self.args.poll_interval, self.args.poll_timeout,
                        session=self.session)
                    # Refresh token after potentially long polling loop.
                    if self.session is not None:
                        self.session.ensure_token()
                    report = download_and_parse_report(
                        self.api, rec, transfer_id, run_dir)
                    # Check system logs for fallback worker evidence.
                    log_evidence = check_fallback_in_logs(
                        self.api, rec, transfer_id)
                    capture["logs_fallback_started"] = log_evidence["fallback_started"]
                    capture["logs_fallback_done"] = log_evidence["fallback_done"]
                    capture["logs_fb_transferred"] = log_evidence["fb_transferred"]
                    capture["logs_fb_failed"] = log_evidence["fb_failed"]
                    if case.direction == "download":
                        self._verify_download_landing(rec, dl_dst, expected_files, report)
                    verdict, reasons = evaluate(
                        case, state, capture, report, expected_files)
        except KeyboardInterrupt:
            verdict, reasons = "INTERRUPTED", ["interrupted by user (Ctrl+C)"]
            rec.add("interrupted", "local", "", ok=False,
                    detail="Ctrl+C received — saving partial steps")
        except Exception as exc:  # noqa: BLE001
            verdict, reasons = "ERROR", [f"{type(exc).__name__}: {exc}"]
            rec.add("exception", "local", "", ok=False, detail=str(exc))

        # Per-case cleanup: empty the bucket + remove generated /bryck data.
        self.cleanup_case(rec, case)

        result = {
            "case_id": case.cid,
            "group": case.group,
            "direction": case.direction,
            "dataset": ds.key,
            "spec": ds.spec,
            "profile": case.profile,
            "fault": {"fail_pct": case.prof.fail_pct,
                      "crash_pct": case.prof.crash_pct,
                      "seed": self.args.seed,
                      "crash_mode": "abort"},
            "hi_perf_opt": case.hi_perf,
            "fallback_enabled": case.fallback_enabled,
            "expect": case.expect,
            "description": case.desc,
            "src": src if case.direction == "upload" else dst,
            "dst": dst if case.direction == "upload" else dl_dst,
            "transfer_id": transfer_id,
            "final_state": state,
            "expected_files": expected_files,
            "capture": capture,
            "report": report,
            "verdict": verdict,
            "reasons": reasons,
            "started": started,
            "finished": _now(),
            "steps": rec.steps,
        }
        _write_json(run_dir / "report.json", result)
        LOG.info("=== %s -> %s (%s)", case.cid, verdict,
                 "; ".join(reasons) or "ok")
        if verdict == "INTERRUPTED":
            raise KeyboardInterrupt
        return result

    def _verify_download_landing(self, rec: Recorder, dl_dst: str,
                                 expected_files: int | None, report: dict) -> None:
        rc, out, _ = self.host.run(
            rec, "verify download landing",
            f"find {dl_dst} -type f 2>/dev/null | wc -l", check=False)
        try:
            landed = int(out.strip().splitlines()[-1]) if out.strip() else 0
        except (ValueError, IndexError):
            landed = None
        report["downloaded_files"] = landed
        detail = f"{landed} file(s) landed under {dl_dst}"
        if expected_files and landed is not None and landed < expected_files:
            detail += f" (< expected {expected_files})"
        rec.add("download landing summary", "local", "", ok=True, detail=detail)

    def cleanup_case(self, rec: Recorder, case: Case) -> None:
        """After a case: empty the object-store bucket and drop generated data."""
        if self.args.skip_cleanup:
            rec.add("skip cleanup", "local", "", detail="--skip-cleanup")
            return
        ds = case.ds
        bucket_root = _bucket_root(self.args.bucket_base)
        aws_cmd = (f"aws s3 rm --recursive {bucket_root} "
                   f"--endpoint-url {self.args.endpoint_url}")
        src = f"{self.args.src_base}/{ds.tier_dir}"
        dl = f"{self.args.dl_base}/{ds.tier_dir}"
        # Guard: never rm a bare base / root path.
        src_rm = f"rm -rf {src}" if ds.tier_dir else "true"
        dl_rm = f"rm -rf {dl}" if ds.tier_dir else "true"
        if self.host.dry_run:
            rec.add("cleanup bucket", "plan", aws_cmd, detail="(dry-run)")
            rec.add("cleanup /bryck source", "plan", src_rm, detail="(dry-run)")
            rec.add("cleanup /bryck download", "plan", dl_rm, detail="(dry-run)")
            return
        self.host.run(rec, "cleanup bucket", aws_cmd, check=False)
        self.host.run(rec, "cleanup /bryck source", src_rm, check=False)
        self.host.run(rec, "cleanup /bryck download", dl_rm, check=False)

    def finalize(self, rec: Recorder) -> None:
        if self.args.keep_config:
            rec.add("keep config", "local", "", detail="--keep-config: leaving fault config in place")
            return
        restore_config(self.host, rec, self.config_backup_remote,
                       self.args.config, self.args.service)


# =============================================================================
# Reporting (combined JSON + HTML)
# =============================================================================

def _write_json(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, indent=2, default=str), encoding="utf-8")


VERDICT_COLORS = {
    "PASS": "#1a7f37", "FAIL": "#cf222e", "WARN": "#9a6700",
    "ERROR": "#82071e", "PLANNED": "#0969da", "INTERRUPTED": "#8250df",
}


def render_html(results: list[dict], meta: dict) -> str:
    def esc(x: Any) -> str:
        return html.escape(str(x))

    counts: dict[str, int] = {}
    for r in results:
        counts[r["verdict"]] = counts.get(r["verdict"], 0) + 1
    summary_badges = " ".join(
        f'<span class="badge" style="background:{VERDICT_COLORS.get(k, "#57606a")}">'
        f'{esc(k)}: {v}</span>'
        for k, v in sorted(counts.items())
    )

    rows = []
    for r in results:
        color = VERDICT_COLORS.get(r["verdict"], "#57606a")
        rows.append(
            f'<tr>'
            f'<td><a href="#{esc(r["case_id"])}">{esc(r["case_id"])}</a></td>'
            f'<td>{esc(r["group"])}</td>'
            f'<td>{esc(r["direction"])}</td>'
            f'<td>{esc(r["dataset"])}</td>'
            f'<td>{esc(r["profile"])} ({r["fault"]["fail_pct"]}/{r["fault"]["crash_pct"]})</td>'
            f'<td>{esc(r["hi_perf_opt"])}</td>'
            f'<td>{esc(r["fallback_enabled"])}</td>'
            f'<td>{esc(r.get("final_state") or "-")}</td>'
            f'<td><span class="badge" style="background:{color}">{esc(r["verdict"])}</span></td>'
            f'</tr>'
        )

    sections = []
    for r in results:
        color = VERDICT_COLORS.get(r["verdict"], "#57606a")
        step_rows = []
        for s in r["steps"]:
            ok = s.get("ok")
            ok_txt = "" if ok is None else ("&#10003;" if ok else "&#10007;")
            ok_col = "#1a7f37" if ok else ("#cf222e" if ok is False else "#57606a")
            out = ""
            if s.get("stdout") or s.get("stderr"):
                stderr_part = ("<br>[stderr] " + esc(s["stderr"])) if s.get("stderr") else ""
                out = f'<pre class="io">{esc(s.get("stdout", ""))}{stderr_part}</pre>'
            cmd = f'<code>{esc(s["command"])}</code>' if s.get("command") else ""
            detail = ("<div class=detail>" + esc(s["detail"]) + "</div>") if s.get("detail") else ""
            rc_txt = "" if s.get("rc") is None else ("rc=" + esc(s["rc"]))
            step_rows.append(
                f'<tr><td>{s["seq"]}</td>'
                f'<td>{esc(s["kind"])}</td>'
                f'<td>{esc(s["name"])}<br>{cmd}{detail}{out}</td>'
                f'<td style="color:{ok_col};font-weight:700">{ok_txt} {rc_txt}</td></tr>'
            )
        reasons = "".join(f"<li>{esc(x)}</li>" for x in r.get("reasons", []))
        rep = r.get("report", {})
        cap = r.get("capture", {})
        metrics = (
            f'transfer_id={esc(r.get("transfer_id"))} · state={esc(r.get("final_state") or "-")} · '
            f'expected_files={esc(r.get("expected_files"))} · '
            f'fallback_started={esc(cap.get("logs_fallback_started", "-"))} · '
            f'retried_rows={esc(rep.get("retried_rows"))} · '
            f'mp_ok={esc(rep.get("mp_ok_rows"))} · '
            f'success_rows={esc(rep.get("success_rows"))} · '
            f'fallback_ok={esc(rep.get("fallback_ok"))} · '
            f'failed_rows={esc(rep.get("failed_rows"))}'
        )
        sections.append(
            f'<section id="{esc(r["case_id"])}" class="case">'
            f'<h3>{esc(r["case_id"])} '
            f'<span class="badge" style="background:{color}">{esc(r["verdict"])}</span></h3>'
            f'<p class="desc">{esc(r["description"])}</p>'
            f'<p class="meta">{metrics}</p>'
            f'{("<ul class=reasons>" + reasons + "</ul>") if reasons else ""}'
            f'<details open><summary>Steps &amp; commands ({len(r["steps"])})</summary>'
            f'<table class="steps"><thead><tr><th>#</th><th>kind</th>'
            f'<th>step / command</th><th>result</th></tr></thead>'
            f'<tbody>{"".join(step_rows)}</tbody></table></details>'
            f'</section>'
        )

    style = """
    body{font-family:-apple-system,Segoe UI,Roboto,Helvetica,Arial,sans-serif;margin:0;background:#f6f8fa;color:#1f2328}
    header{background:#24292f;color:#fff;padding:18px 28px}
    header h1{margin:0;font-size:20px}
    .wrap{padding:20px 28px}
    .badge{color:#fff;padding:2px 8px;border-radius:10px;font-size:12px;font-weight:700}
    table{border-collapse:collapse;width:100%;background:#fff;margin:12px 0;font-size:13px}
    th,td{border:1px solid #d0d7de;padding:6px 8px;text-align:left;vertical-align:top}
    th{background:#eaeef2}
    .summary td,.summary th{white-space:nowrap}
    .case{background:#fff;border:1px solid #d0d7de;border-radius:8px;padding:14px 18px;margin:16px 0}
    .case h3{margin:0 0 6px}
    .desc{color:#57606a;margin:4px 0}
    .meta{font-family:ui-monospace,SFMono-Regular,Consolas,monospace;font-size:12px;color:#24292f;background:#f6f8fa;padding:6px 8px;border-radius:6px}
    .reasons{margin:8px 0;color:#9a6700}
    code{font-family:ui-monospace,SFMono-Regular,Consolas,monospace;font-size:12px;color:#0550ae;word-break:break-all}
    .detail{color:#57606a;font-size:12px;margin-top:2px}
    pre.io{background:#0d1117;color:#c9d1d9;padding:8px;border-radius:6px;overflow:auto;max-height:220px;font-size:11px;margin:6px 0 0}
    table.steps td:first-child{width:32px;text-align:right;color:#57606a}
    summary{cursor:pointer;font-weight:600;margin:6px 0}
    """
    head = (
        f'<tr><th>case</th><th>group</th><th>dir</th><th>dataset</th>'
        f'<th>profile (F/C%)</th><th>HI_PERF</th><th>FALLBACK</th>'
        f'<th>state</th><th>verdict</th></tr>'
    )
    return (
        f'<!doctype html><html><head><meta charset="utf-8">'
        f'<title>CloudCp Fallback Test Report</title><style>{style}</style></head><body>'
        f'<header><h1>CloudCp Fallback Test Report</h1>'
        f'<div>{esc(meta.get("generated"))} · host {esc(meta.get("host"))} · '
        f'seed {esc(meta.get("seed"))} · {len(results)} case(s)</div></header>'
        f'<div class="wrap"><p>{summary_badges}</p>'
        f'<table class="summary"><thead>{head}</thead><tbody>{"".join(rows)}</tbody></table>'
        f'{"".join(sections)}</div></body></html>'
    )


# =============================================================================
# CLI
# =============================================================================

def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    p = argparse.ArgumentParser(
        description="CloudCp transfer-fallback test harness.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    sel = p.add_argument_group("selection")
    sel.add_argument("--all", action="store_true", help="Run every case (incl. negatives).")
    sel.add_argument("--one", help="Run one case, or a comma-separated list, by # index or case id.")
    sel.add_argument("--from", dest="from_id", help="Start case (# index or id, inclusive) in catalog order.")
    sel.add_argument("--to", dest="to_id", help="End case (# index or id, inclusive) in catalog order.")
    sel.add_argument("--negative", action="store_true", help="Run only the FB-N-* negative cases.")
    sel.add_argument("--negative-case", help="Run one/comma-separated negative case(s) by # index or id.")
    sel.add_argument("--list", action="store_true", help="List all cases and exit.")

    ex = p.add_argument_group("execution")
    ex.add_argument("--dry-run", action="store_true",
                    help="Print the plan (commands + API calls) without executing.")
    ex.add_argument("--manual", action="store_true",
                    help="Interactive mode: print each case (with F/C fault %%) and "
                         "prompt to execute or skip it.")
    ex.add_argument("--skip-datagen", action="store_true", help="Reuse already-materialised data.")
    ex.add_argument("--skip-seed", action="store_true",
                    help="For download cases, reuse existing bucket objects (no seeding upload).")
    ex.add_argument("--keep-config", action="store_true",
                    help="Do not restore the original config after the run.")
    ex.add_argument("--skip-cleanup", action="store_true",
                    help="Do not empty the bucket / remove generated /bryck data after each case.")
    ex.add_argument("--seed", type=int, default=DEFAULT_FAULT_SEED,
                    help=f"FAULT_SEED value (default {DEFAULT_FAULT_SEED}).")
    ex.add_argument("--poll-interval", type=int, default=DEF_POLL_INTERVAL)
    ex.add_argument("--poll-timeout", type=int, default=DEF_POLL_TIMEOUT)
    ex.add_argument("--verbose", action="store_true")

    paths = p.add_argument_group("paths / hosts")
    paths.add_argument("--cli-dir", default=str(DEFAULT_CLI_DIR),
                       help="bryckclient-cli directory (login.json / cloud_ops.json / modules).")
    paths.add_argument("--spec-dir", default=str(DEFAULT_SPEC_DIR))
    paths.add_argument("--out-dir", default=str(DEFAULT_OUT_DIR))
    paths.add_argument("--login", default=None, help="Path to login.json (default: <cli-dir>/login.json).")
    paths.add_argument("--cloud-ops", default=None,
                       help="Path to cloud_ops.json (default: <cli-dir>/cloud_ops.json).")
    paths.add_argument("--datagen", default=DEF_DATAGEN)
    paths.add_argument("--config", default=DEF_CONFIG)
    paths.add_argument("--service", default=DEF_SERVICE)
    paths.add_argument("--transfer-logs-dir", default=DEF_TRANSFER_LOGS)
    paths.add_argument("--endpoint-url", default=DEF_ENDPOINT)
    paths.add_argument("--src-base", default=DEF_SRC_BASE)
    paths.add_argument("--bucket-base", default=DEF_BUCKET_BASE)
    paths.add_argument("--dl-base", default=DEF_DL_BASE)

    # -- component-level fallback tests (fallback_worker + mp_batch_retry) -----
    # These run the two internal fallback mechanisms in isolation via SSH,
    # staging their exact on-disk inputs (see plan_cp_component_fallback.md).
    comp = p.add_argument_group("component tests (internal fallback mechanisms)")
    comp.add_argument("--component", action="store_true",
                      help="Run the component fallback suite (fallback_worker + mp_batch_retry).")
    comp.add_argument("--component-one",
                      help="Run one/comma-separated component case(s) by # index or id.")
    comp.add_argument("--component-negative", action="store_true",
                      help="Run only the component negative cases.")
    comp.add_argument("--component-list", action="store_true",
                      help="List all component cases and exit.")
    comp.add_argument("--heavy", action="store_true",
                      help="Include heavy datasets (large, scale) in the component suite.")
    comp.add_argument("--component-bucket", default="omicron",
                      help="Destination bucket for both component mechanisms (default: omicron).")
    comp.add_argument("--region", default="us-west-1",
                      help="AWS region passed to retry_whole_batch (default: us-west-1).")
    comp.add_argument("--venv-python", default="/opt/bryck/.venv/bryck/bin/python3",
                      help="Target interpreter that imports bryckcloud.")
    comp.add_argument("--batchmeta-dir",
                      default="/opt/bryck/bryckapi/downloads/bcloud_batchmeta",
                      help="BATCH_FILE_DIR (batch-meta root) on the target.")
    comp.add_argument("--pool-size", type=int, default=16,
                      help="Fallback worker --pool-size (default: 16).")
    return p.parse_args(argv)


def _resolve_ids(tokens: list[str], catalog: list[Case]) -> list[str]:
    """Map selection tokens to case ids, preserving order.

    A token may be a case id (FB-U-01, case-insensitive) or a 1-based catalog
    index as shown in the '#' column of --list (e.g. 1, 22). Unknown tokens are
    reported on stderr and skipped.
    """
    order = [c.cid for c in catalog]
    lower = {cid.lower(): cid for cid in order}
    resolved: list[str] = []
    unknown: list[str] = []
    for tok in tokens:
        if tok.lower() in lower:
            resolved.append(lower[tok.lower()])
        elif tok.isdigit() and 1 <= int(tok) <= len(order):
            resolved.append(order[int(tok) - 1])
        else:
            unknown.append(tok)
    if unknown:
        print(f"Unknown case id/index: {', '.join(unknown)}. "
              f"Run --list to see valid ids (1..{len(order)} or FB-U-01).",
              file=sys.stderr)
    return resolved


def select_cases(args, catalog: list[Case]) -> list[Case]:
    by_id = {c.cid: c for c in catalog}
    order = [c.cid for c in catalog]

    if args.negative and not (args.one or args.from_id or args.negative_case):
        return [c for c in catalog if c.group == "NEGATIVE"]
    if args.negative_case:
        tokens = [x.strip() for x in args.negative_case.split(",") if x.strip()]
        return [by_id[i] for i in _resolve_ids(tokens, catalog)]
    if args.one:
        tokens = [x.strip() for x in args.one.split(",") if x.strip()]
        return [by_id[i] for i in _resolve_ids(tokens, catalog)]
    if args.from_id or args.to_id:
        from_id = _resolve_ids([args.from_id], catalog) if args.from_id else []
        to_id = _resolve_ids([args.to_id], catalog) if args.to_id else []
        start = order.index(from_id[0]) if from_id else 0
        end = order.index(to_id[0]) if to_id else len(order) - 1
        if start > end:
            start, end = end, start
        return [by_id[cid] for cid in order[start:end + 1]]
    if args.all:
        return list(catalog)
    return []


def _prompt_manual(case: Case) -> str:
    """Print a case (with F/C fault %) and prompt to execute/skip/quit.

    Returns one of: "execute", "skip", "quit".
    """
    prof = case.prof
    print("\n" + "=" * 66)
    print(f"  {case.cid}  [{case.group}]  {case.direction.upper()}")
    print("-" * 66)
    print(f"  dataset          : {case.ds.key} ({case.ds.spec})")
    print(f"  profile          : {case.profile} — {prof.desc}")
    print(f"  FAIL %  (F)      : {prof.fail_pct}%")
    print(f"  CRASH % (C)      : {prof.crash_pct}%")
    print(f"  HI_PERF_OPT      : {case.hi_perf}")
    print(f"  FALLBACK_ENABLED : {case.fallback_enabled}")
    print(f"  expectation      : {case.expect}")
    print(f"  description      : {case.desc}")
    print("=" * 66)
    while True:
        try:
            ans = input("  [e]xecute / [s]kip / [q]uit ? ").strip().lower()
        except EOFError:
            return "quit"
        if ans in ("e", "execute", "y", "yes"):
            return "execute"
        if ans in ("s", "skip", "n", "no"):
            return "skip"
        if ans in ("q", "quit"):
            return "quit"
        print("  Please answer 'e' (execute), 's' (skip), or 'q' (quit).")


def _load_json(path: Path) -> dict:
    return json.loads(path.read_text(encoding="utf-8"))


def _load_creds(cloud_ops: Path) -> dict:
    cfg = _load_json(cloud_ops)
    ct = str(cfg.get("cloud_type", "aws")).lower()
    return {
        "cloud_type": ct,
        "access_key_id": cfg.get("access_key_id"),
        "secret_access_key": cfg.get("secret_access_key"),
        "region": cfg.get("region") or "us-east-1",
    }


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv)
    logging.basicConfig(
        level=logging.DEBUG if args.verbose else logging.INFO,
        format="%(asctime)s %(levelname)s %(name)s: %(message)s",
    )
    catalog = build_catalog()

    if args.list:
        print(f"{'#':<4} {'CASE':<9} {'GROUP':<10} {'DIR':<8} {'DATASET':<8} "
              f"{'PROF':<4} {'HP':<5} {'FB':<5} EXPECT")
        for i, c in enumerate(catalog, 1):
            print(f"{i:<4} {c.cid:<9} {c.group:<10} {c.direction:<8} {c.dataset:<8} "
                  f"{c.profile:<4} {str(c.hi_perf):<5} {str(c.fallback_enabled):<5} {c.expect}")
        print(f"\n{len(catalog)} cases. Select by # (1..{len(catalog)}) or case id. "
              "Profiles: "
              + ", ".join(f"{k}={v.fail_pct}/{v.crash_pct}" for k, v in PROFILES.items()))
        return 0

    if args.component_list:
        from cloudcp_component_fallback_test import print_component_list
        print_component_list()
        return 0

    component_mode = bool(args.component or args.component_one or args.component_negative)

    selected = select_cases(args, catalog)
    if not selected and not component_mode:
        print("No cases selected. Use --all, --one, --from/--to, --negative, "
              "--component, --component-one, --component-negative, or --list.",
              file=sys.stderr)
        return 2

    cli_dir = Path(args.cli_dir)
    login = Path(args.login) if args.login else cli_dir / "login.json"
    cloud_ops = Path(args.cloud_ops) if args.cloud_ops else cli_dir / "cloud_ops.json"
    out_dir = Path(args.out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    api = None
    host = RemoteHost(None, dry_run=True)
    creds = {"cloud_type": "aws", "access_key_id": None,
             "secret_access_key": None, "region": "us-east-1"}
    session = None
    host_ip = "(dry-run)"

    if not args.dry_run:
        # Import the bryckclient-cli modules lazily so --list / --dry-run work
        # without paramiko/requests installed.
        sys.path.insert(0, str(cli_dir))
        try:
            from session import ApiSession           # type: ignore
            from bryck_api import BryckApi            # type: ignore
            from ssh_runner import SshRunner          # type: ignore
        except ImportError as exc:
            print(f"Failed to import bryckclient-cli modules from {cli_dir}: {exc}",
                  file=sys.stderr)
            return 3
        if not login.is_file() or not cloud_ops.is_file():
            print(f"Missing login.json ({login}) or cloud_ops.json ({cloud_ops}).",
                  file=sys.stderr)
            return 3
        creds = _load_creds(cloud_ops)
        session = ApiSession.from_login_json(str(login))
        session.login()
        api = BryckApi(session)
        ssh = SshRunner.from_session(session)
        ssh.connect()
        host = RemoteHost(ssh, dry_run=False)
        host_ip = session.host

    if component_mode:
        # Delegate to the component-level fallback suite, then tidy up the
        # shared SSH/API session before returning.
        try:
            from cloudcp_component_fallback_test import run_component_suite
            return run_component_suite(args, host, session)
        finally:
            if session is not None:
                try:
                    session.close()
                except Exception:  # noqa: BLE001
                    pass
            if host.ssh is not None:
                try:
                    host.ssh.close()
                except Exception:  # noqa: BLE001
                    pass

    runner = Runner(args, api, host, creds, session=session)
    results: list[dict] = []
    interrupted = False
    skipped: list[str] = []
    try:
        for case in selected:
            if args.manual:
                decision = _prompt_manual(case)
                if decision == "quit":
                    LOG.info("Manual mode: quitting — remaining cases not run.")
                    break
                if decision == "skip":
                    skipped.append(case.cid)
                    LOG.info("Manual mode: skipped %s", case.cid)
                    continue
            result = runner.run_case(case)
            results.append(result)
    except KeyboardInterrupt:
        # run_case catches the interrupt, saves partial steps, appends nothing
        # yet — but re-raises. The last result (INTERRUPTED) was written to
        # its per-case report.json but not yet in results, so grab it.
        last_json = out_dir / case.cid / "report.json"
        if last_json.is_file():
            try:
                results.append(json.loads(last_json.read_text(encoding="utf-8")))
            except Exception:  # noqa: BLE001
                pass
        interrupted = True
        LOG.warning("Interrupted (Ctrl+C) — writing report for %d case(s) "
                    "(including partial).", len(results))
    finally:
        fin_rec = Recorder("finalize")
        try:
            runner.finalize(fin_rec)
        finally:
            _write_json(out_dir / "finalize.json", {"steps": fin_rec.steps})
        if session is not None:
            try:
                session.close()
            except Exception:  # noqa: BLE001
                pass
        if host.ssh is not None:
            try:
                host.ssh.close()
            except Exception:  # noqa: BLE001
                pass

    meta = {"generated": _now(), "host": host_ip, "seed": args.seed,
            "dry_run": args.dry_run, "interrupted": interrupted}
    combined = {"meta": meta, "results": results}
    _write_json(out_dir / "fallback_report.json", combined)
    (out_dir / "fallback_report.html").write_text(
        render_html(results, meta), encoding="utf-8")

    tally: dict[str, int] = {}
    for r in results:
        tally[r["verdict"]] = tally.get(r["verdict"], 0) + 1
    if skipped:
        LOG.info("Manual mode skipped %d case(s): %s",
                 len(skipped), ", ".join(skipped))
    if interrupted:
        LOG.info("Partial run (interrupted). %s",
                 " ".join(f"{k}={v}" for k, v in sorted(tally.items())))
    else:
        LOG.info("Done. %s", " ".join(f"{k}={v}" for k, v in sorted(tally.items())))
    LOG.info("Reports: %s , %s",
             out_dir / "fallback_report.json", out_dir / "fallback_report.html")

    # Non-zero exit if anything failed/errored (useful for CI).
    bad = tally.get("FAIL", 0) + tally.get("ERROR", 0)
    return 1 if bad or interrupted else 0


if __name__ == "__main__":
    sys.exit(main())
